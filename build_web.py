# -*- coding: utf-8 -*-
"""Ghep web_template.html + optimizer.js + diem cau thu (tu xlsx) -> index.html (trang co nut bam)."""
import json, os
from openpyxl import load_workbook

BASE="/Users/tuyennd/Documents/VN_NAMI/Tools/GenPlayer/"

# 1) doc diem 20 cau thu tu xlsx
LOWPRIO={"Duy Hùng","Quý Tàu","Trịnh Mạnh","Mạnh Tiến","Quang Minh"}   # uu tien du bi
wb=load_workbook(BASE+"BangDanhGia_CauThu.xlsx",data_only=True); ws=wb["ChamDiem"]
HR=next(r for r in range(1,40) if ws.cell(r,1).value=="STT"); FIRST=HR+1
players=[]
for idx in range(20):
    r=FIRST+3*idx
    nm=ws.cell(r,2).value
    KT,CH,DD,PN,TC,TL,DN,TD,TM=[ws.cell(r,c).value for c in range(8,17)]
    players.append({"n":nm,"KT":KT,"CH":CH,"DD":DD,"PN":PN,
                    "TC":TC,"TL":TL,"DN":DN,"TD":TD,"TM":TM,"lp":nm in LOWPRIO})
players_js=json.dumps(players,ensure_ascii=False)

# 2) doc optimizer.js, bo phan test node
opt=open(BASE+"optimizer.js",encoding="utf-8").read()
opt=opt.split("// ===== Node test =====")[0].rstrip()

# 3) firebase config (neu co file firebase_config.json) -> bat dong bo real-time
fb="null"
if os.path.exists(BASE+"firebase_config.json"):
    fb=open(BASE+"firebase_config.json",encoding="utf-8").read().strip()

# 4) nhet vao template
tpl=open(BASE+"web_template.html",encoding="utf-8").read()
html=(tpl.replace("/*__OPTIMIZER__*/",opt)
         .replace("/*__PLAYERS__*/",players_js)
         .replace("/*__FIREBASE__*/",fb))

open(BASE+"index.html","w",encoding="utf-8").write(html)
print("Đã tạo index.html ·",len(players),"cầu thủ · Firebase:",("BẬT" if fb!="null" else "tắt (cục bộ)"),"·",len(html),"ký tự")
assert "/*__OPTIMIZER__*/" not in html and "/*__PLAYERS__*/" not in html and "/*__FIREBASE__*/" not in html, "Còn placeholder!"
print("OK: không còn placeholder")
