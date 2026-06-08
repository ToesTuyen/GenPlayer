# -*- coding: utf-8 -*-
"""Quy doi tra loi Tuyen (tier + vi tri + toc do + do nhiet) -> 9 tieu chi + vi tri 6 vai."""
from openpyxl import load_workbook

# (ten, tier, coarse positions, speed: F/S/_, workrate: H/L/_)
# tier: NS/K/TB/Y ; speed F=nhanh S=cham _=tb ; wr H=chiu chay L=luoi _=tb
DATA = [
    ("Cris Dao",          "TB", {"HV","TV"},      "F","L"),
    ("Doan Ha",           "TB", {"TĐ"},           "S","H"),
    ("Duy Hùng",          "Y",  {"HV","TV"},      "S","_"),
    ("Hoàng Duy",         "K",  {"HV","TV"},      "F","H"),
    ("Hoàng Thiên",       "NS", {"GK","TĐ"},      "F","H"),
    ("Hưng",              "K",  {"HV","TV"},      "S","H"),
    ("Huy Hoàng",         "NS", {"HV","TV","TĐ"}, "F","H"),
    ("Mạnh Khắc",         "TB", {"HV","TV","TĐ"}, "S","L"),
    ("Mạnh Tiến",         "TB", {"GK","HV"},      "F","H"),
    ("Nam Khắc",          "K",  {"HV"},           "F","H"),   # da hau ve thoi
    ("Nguyễn Danh Tuyên", "K",  {"HV","TV"},      "F","H"),
    ("Nguyễn Khắc Trọng", "TB", {"GK","HV"},      "S","H"),
    ("Nguyen Quy Bong",   "NS", {"TV","TĐ"},      "F","H"),
    ("Nguyễn Văn Quyền",  "Y",  {"HV","TV"},      "S","L"),
    ("Quang Minh",        "TB", {"GK","HV","TV"}, "F","H"),
    ("Quý Tàu",           "Y",  {"HV","TV"},      "S","L"),
    ("Quyết Nguyễn",      "K",  {"HV","TV","TĐ"}, "F","H"),
    ("Trịnh Mạnh",        "TB", {"HV","TV"},      "S","H"),
    ("Trịnh Tuấn Anh",    "K",  {"HV","TĐ"},      "F","H"),
    ("Văn Thanh",         "NS", {"TĐ","HV"},      "S","H"),
]

# Chinh tay tung chi so (ghi de sau khi tinh tu tier+vi tri). Key chi so: KT,CH,DD,PN,TC,TL,DN,TD,TM
OVERRIDES = {
    "Doan Ha": {"TL": 7},     # tang the luc 1 chut (do nhiet ĐN giu nguyen 8)
}

BASE = {"NS":8.5,"K":7.0,"TB":5.5,"Y":4.0}
TIER_LB = {"NS":"Ngôi sao","K":"Khá","TB":"Trung bình","Y":"Yếu"}
TM_TIER = {"NS":7.5,"K":6.0,"TB":5.0,"Y":4.0}
TILT = {  # [KT, CH, DĐ, PN, TL, TD]
    "HV":[-0.5,0.0,-1.5,1.5,0.5,0.5],
    "TV":[ 0.5,1.0,-0.5,0.0,0.5,1.0],
    "TĐ":[ 1.0,-0.5,1.5,-1.5,0.5,0.0],
}
SPD = {"F":8.0,"S":3.5,"_":5.5}
WR  = {"H":8.0,"L":3.0,"_":5.5}

def rhu(x): return max(1,min(10,int(x+0.5)))

def compute(tier, pos, spd, wr):
    base=BASE[tier]
    outp=[p for p in pos if p in TILT]
    avg=[sum(TILT[p][i] for p in outp)/len(outp) for i in range(6)] if outp else [0]*6
    KT,CH,DD,PN,TL,TD = [rhu(base+avg[i]) for i in range(6)]
    TC = rhu(SPD[spd] + 0.3*(base-5.5))
    DN = rhu(WR[wr]  + 0.2*(base-5.5))
    TM = rhu(TM_TIER[tier]) if "GK" in pos else 0
    sc = dict(KT=KT,CH=CH,DD=DD,PN=PN,TC=TC,TL=TL,DN=DN,TD=TD,TM=TM)

    # vi tri 6 vai
    playable=set()
    if "HV" in pos: playable |= {"Thòng","HV"}
    if "TV" in pos: playable |= {"TVtt","TV"}
    if "TĐ" in pos: playable |= {"TĐ"}
    if "GK" in pos: playable |= {"GK"}
    fit={
        "GK":   TM,
        "Thòng":0.40*PN+0.25*TD+0.20*TL+0.15*CH - 0.20*max(0,DD-6),
        "HV":   0.30*PN+0.25*TC+0.20*TL+0.15*DN+0.10*TD,
        "TVtt": 0.30*CH+0.30*TD+0.20*KT+0.10*DN+0.10*PN,
        "TV":   0.30*KT+0.20*TC+0.20*CH+0.15*DN+0.15*DD,
        "TĐ":   0.45*DD+0.20*KT+0.15*TL+0.10*TD+0.10*TC,
    }
    atk={"TĐ":5,"TV":4,"TVtt":3,"HV":2,"Thòng":1,"GK":0}  # hoa -> uu tien vai tan cong
    main=max(playable,key=lambda p:(fit[p], atk[p]))
    order_ex=["Thòng","HV","TVtt","TV","TĐ"]              # GK da the hien qua TM>0
    extras=[p for p in order_ex if p in playable and p!=main and fit[p]>=fit[main]-1.5]
    return sc, main, "/".join(extras), fit

path="/Users/tuyennd/Documents/VN_NAMI/Tools/GenPlayer/BangDanhGia_CauThu.xlsx"
wb=load_workbook(path); ws=wb["ChamDiem"]
# tim header row
HR=next(r for r in range(1,40) if ws.cell(r,1).value=="STT")
FIRST=HR+1
SPDLB={"F":"nhanh","S":"chậm","_":""}; WRLB={"H":"chịu chạy","L":"lười","_":""}

print(f"{'#':>2} {'Tên':<19}{'Chính':<6}{'Thêm':<10}KT CH DĐ PN TC TL ĐN TD TM")
print("-"*74)
for idx,(name,tier,pos,spd,wr) in enumerate(DATA):
    sc,main,extra,fit=compute(tier,pos,spd,wr)
    for k,v in OVERRIDES.get(name,{}).items(): sc[k]=v   # ghi de chinh tay
    row=FIRST+3*idx
    ws.cell(row,4,main); ws.cell(row,5,extra)
    vals=[sc["KT"],sc["CH"],sc["DD"],sc["PN"],sc["TC"],sc["TL"],sc["DN"],sc["TD"],sc["TM"]]
    for j,v in enumerate(vals): ws.cell(row,8+j,v)
    tags=", ".join(t for t in (SPDLB[spd],WRLB[wr]) if t)
    ws.cell(row,17,f"Tuyên: {TIER_LB[tier]}{', '+tags if tags else ''}")
    print(f"{idx+1:>2} {name:<19}{main:<6}{extra:<10}" + " ".join(f"{v:>2}" for v in vals))

wb.save(path)
print("\nĐã ghi dòng Tuyên (9 tiêu chí + vị trí 6 vai) vào file.")
