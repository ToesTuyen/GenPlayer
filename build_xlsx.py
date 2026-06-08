# -*- coding: utf-8 -*-
"""File Excel cham diem: 9 tieu chi + 6 vi tri, huong dan + bang 3 dong/player + tab trung binh."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

PLAYERS = [
    "Cris Dao", "Doan Ha", "Duy Hùng", "Hoàng Duy", "Hoàng Thiên",
    "Hưng", "Huy Hoàng", "Mạnh Khắc", "Mạnh Tiến", "Nam Khắc",
    "Nguyễn Danh Tuyên", "Nguyễn Khắc Trọng", "Nguyen Quy Bong",
    "Nguyễn Văn Quyền", "Quang Minh", "Quý Tàu", "Quyết Nguyễn",
    "Trịnh Mạnh", "Trịnh Tuấn Anh", "Văn Thanh",
]
RATERS = ["Tuyên", "Mạnh", "Nam"]

NAVY="1F3864"; BLUE="2E5496"; LBLUE="D6E0F0"; GREY="F2F2F2"; GREY2="E9EDF4"
GREEN="548235"; LGREEN="E2EFDA"; GOLD="FFF2CC"; WHITE="FFFFFF"; DIS="DDDDDD"
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="BFBFBF")
B = Border(left=thin, right=thin, top=thin, bottom=thin)
wrapL = Alignment(horizontal="left", vertical="center", wrap_text=True)
ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# ===================== TAB 1: CHAM DIEM =====================
ws = wb.active; ws.title = "ChamDiem"; ws.sheet_view.showGridLines = False
NCOLS=17; LAST=get_column_letter(NCOLS)
def cmerge(r, text, *, fillc=None, fontc="000000", bold=False, size=11, align=wrapL, height=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    cell=ws.cell(r,1,text); cell.font=Font(bold=bold,color=fontc,size=size); cell.alignment=align
    if fillc:
        for col in range(1,NCOLS+1): ws.cell(r,col).fill=fill(fillc)
    if height: ws.row_dimensions[r].height=height

r=1
cmerge(r, "BẢNG CHẤM ĐIỂM CẦU THỦ — CHIA 2 ĐỘI SÂN 7   (3 người chấm: Tuyên / Mạnh / Nam)",
       fillc=NAVY, fontc=WHITE, bold=True, size=13, align=ctr, height=24); r+=1
cmerge(r, "Mỗi cầu thủ có 3 dòng (Tuyên / Mạnh / Nam). Bấm ô → SỔ XUỐNG chọn. Vị trí / Chân / Mức độ chỉ điền 1 LẦN ở DÒNG ĐẦU (Tuyên). "
          "Chấm SO VỚI MẶT BẰNG CỦA HỘI; khâu nào không rõ để 5–6; TM = 0 nếu không bắt gôn.",
       fillc=GREY, bold=True, height=44); r+=1
cmerge(r, "Ý NGHĨA 9 TIÊU CHÍ", fillc=BLUE, fontc=WHITE, bold=True, align=ctr, height=18); r+=1
cmerge(r, "KT = Kỹ thuật   ·   CH = Chuyền   ·   DĐ = Dứt điểm   ·   PN = Phòng ngự   ·   TC = Tốc độ",
       fillc=LBLUE, height=20); r+=1
cmerge(r, "TL = Thể lực (sức bền/mạnh)   ·   ĐN = Độ nhiệt (chịu chạy, máu lửa)   ·   TD = Tư duy   ·   TM = Thủ môn (0 nếu không bắt gôn)",
       fillc=LBLUE, height=20); r+=1
cmerge(r, "THANG ĐIỂM (chấm so với mặt bằng của hội)", fillc=GREEN, fontc=WHITE, bold=True, align=ctr, height=18); r+=1
cmerge(r, "1–2 Rất yếu   ·   3–4 Dưới trung bình   ·   5–6 Trung bình (để mức này nếu không rõ)   ·   7–8 Khá tốt   ·   9–10 Xuất sắc",
       fillc=LGREEN, height=20); r+=1
cmerge(r, "VỊ TRÍ (6 vai): GK Thủ môn · Thòng (trung vệ giữa) · HV Hậu vệ biên · TVtt Tiền vệ trung tâm · TV Tiền vệ cánh · TĐ Tiền đạo.   "
          "'Vị trí chính' = sở trường; 'Đá được thêm' = vai trò khác chơi được.",
       fillc=GOLD, height=32); r+=1
r+=1

HEADER_ROW=r
HEAD=[
    "STT","Tên cầu thủ","Người chấm","Vị trí chính","Đá được\nthêm","Chân\nthuận","Mức độ\nhay đi đá",
    "Kỹ thuật\n(KT)","Chuyền\n(CH)","Dứt điểm\n(DĐ)","Phòng ngự\n(PN)","Tốc độ\n(TC)",
    "Thể lực\n(TL)","Độ nhiệt\n(ĐN)","Tư duy\n(TD)","Thủ môn\n(TM·0-10)","Ghi chú",
]
for i,h in enumerate(HEAD, start=1):
    c=ws.cell(HEADER_ROW,i,h); c.font=Font(bold=True,color=WHITE); c.fill=fill(NAVY); c.alignment=ctr; c.border=B
ws.row_dimensions[HEADER_ROW].height=42

first=HEADER_ROW+1
for idx, name in enumerate(PLAYERS):
    r1=first+3*idx
    block_fill = WHITE if idx%2==0 else GREY2
    for k, rater in enumerate(RATERS):
        rr=r1+k
        ws.cell(rr,1, idx+1 if k==0 else None).alignment=ctr
        ws.cell(rr,2, name if k==0 else None).alignment=wrapL
        ws.cell(rr,3, rater).alignment=ctr
        for col in range(1,NCOLS+1):
            cell=ws.cell(rr,col); cell.border=B; cell.fill=fill(block_fill)
            if col>=3: cell.alignment=ctr
        if k>0:
            for col in (4,5,6,7): ws.cell(rr,col).fill=fill(DIS)
    ws.cell(r1,2).font=Font(bold=True)
last=first+3*len(PLAYERS)-1

def add_dv(formula, col, r_from, r_to):
    dv=DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.prompt="Sổ xuống và chọn"; dv.promptTitle="Chọn giá trị"
    ws.add_data_validation(dv); L=get_column_letter(col); dv.add(f"{L}{r_from}:{L}{r_to}")
for idx in range(len(PLAYERS)):
    r1=first+3*idx
    add_dv('"GK,Thòng,HV,TVtt,TV,TĐ"',4,r1,r1)         # Vi tri chinh (6)
    add_dv('"Trái,Phải,Cả hai"',6,r1,r1)
    add_dv('"Thường xuyên,Thỉnh thoảng,Hiếm khi"',7,r1,r1)
score='"1,2,3,4,5,6,7,8,9,10"'
for col in range(8,16): add_dv(score,col,first,last)    # KT..TD (8..15)
add_dv('"0,1,2,3,4,5,6,7,8,9,10"',16,first,last)        # TM

widths={1:5,2:17,3:10,4:11,5:11,6:8,7:10,8:8,9:8,10:8,11:9,12:8,13:8,14:8,15:8,16:9,17:20}
for col,w in widths.items(): ws.column_dimensions[get_column_letter(col)].width=w
ws.auto_filter.ref=f"A{HEADER_ROW}:{LAST}{last}"
ws.freeze_panes=f"D{first}"

# ===================== TAB 2: TONG HOP =====================
s = wb.create_sheet("TongHop"); s.sheet_view.showGridLines = False
SHEAD=["STT","Tên cầu thủ","Vị trí\nchính","KT","CH","DĐ","PN","TC","TL","ĐN","TD","TM","Overall\n(TB 8 mục)"]
SNC=len(SHEAD); SLC=get_column_letter(SNC)
s.merge_cells(start_row=1,start_column=1,end_row=1,end_column=SNC)
t=s.cell(1,1,"TỔNG HỢP — ĐIỂM TRUNG BÌNH 3 NGƯỜI CHẤM (Tuyên/Mạnh/Nam) — tự động từ tab 'ChamDiem'")
t.font=Font(bold=True,color=WHITE,size=12); t.alignment=ctr
for col in range(1,SNC+1): s.cell(1,col).fill=fill(NAVY)
s.row_dimensions[1].height=24
SH=3
for i,h in enumerate(SHEAD,start=1):
    c=s.cell(SH,i,h); c.font=Font(bold=True,color=WHITE); c.fill=fill(BLUE); c.alignment=ctr; c.border=B
s.row_dimensions[SH].height=30
sfirst=SH+1
for idx,name in enumerate(PLAYERS):
    rr=sfirst+idx; r1=first+3*idx; r3=r1+2
    s.cell(rr,1,idx+1).alignment=ctr
    s.cell(rr,2,name).alignment=wrapL
    s.cell(rr,3,f"=IF(ChamDiem!D{r1}=\"\",\"\",ChamDiem!D{r1})").alignment=ctr
    for j in range(9):  # KT..TM <- ChamDiem cols 8..16
        scol=4+j; ccol=get_column_letter(8+j)
        s.cell(rr,scol,f"=IFERROR(ROUND(AVERAGE(ChamDiem!{ccol}{r1}:{ccol}{r3}),1),\"\")").alignment=ctr
    # Overall = TB cua 8 muc ngoai san (KT..TD = cols 4..11)
    s.cell(rr,13,f"=IFERROR(ROUND(AVERAGE(D{rr}:K{rr}),1),\"\")").alignment=ctr
    s.cell(rr,13).font=Font(bold=True)
    for col in range(1,SNC+1):
        s.cell(rr,col).border=B
        if idx%2==1: s.cell(rr,col).fill=fill(GREY)
slast=sfirst+len(PLAYERS)-1
rule=ColorScaleRule(start_type='num',start_value=1,start_color='F8696B',
                    mid_type='num',mid_value=5.5,mid_color='FFEB84',
                    end_type='num',end_value=10,end_color='63BE7B')
s.conditional_formatting.add(f"D{sfirst}:M{slast}", rule)
swidths={1:5,2:18,3:8,4:6,5:6,6:6,7:6,8:6,9:6,10:6,11:6,12:6,13:12}
for col,w in swidths.items(): s.column_dimensions[get_column_letter(col)].width=w
s.auto_filter.ref=f"A{SH}:{SLC}{slast}"
s.freeze_panes=f"A{sfirst}"

wb.active = 0
out="/Users/tuyennd/Documents/VN_NAMI/Tools/GenPlayer/BangDanhGia_CauThu.xlsx"
wb.save(out)
print("Saved:", out, "| Tabs:", wb.sheetnames)
print("ChamDiem: header", HEADER_ROW, "data", first, "-", last, "| cols", NCOLS)
print("TongHop: data", sfirst, "-", slast)
