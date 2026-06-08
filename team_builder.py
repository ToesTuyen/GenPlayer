# -*- coding: utf-8 -*-
"""Chia 20 cau thu -> 2 doi, doi hinh 1-3-3 + 3 du bi, can ca chinh lan du bi.
Duyet toan bo C(19,9)=92378 cach chia, moi doi gan vi tri toi uu bang Hungarian."""
import numpy as np
from itertools import combinations
from scipy.optimize import linear_sum_assignment
from openpyxl import load_workbook

path="/Users/tuyennd/Documents/VN_NAMI/Tools/GenPlayer/BangDanhGia_CauThu.xlsx"
wb=load_workbook(path, data_only=True); ws=wb["ChamDiem"]
HR=next(r for r in range(1,40) if ws.cell(r,1).value=="STT"); FIRST=HR+1
N=20
names=[]; S=[]  # S[i] = dict 9 chi so
for idx in range(N):
    r=FIRST+3*idx
    names.append(ws.cell(r,2).value)
    KT,CH,DD,PN,TC,TL,DN,TD,TM=[ws.cell(r,c).value for c in range(8,17)]
    S.append(dict(KT=KT,CH=CH,DD=DD,PN=PN,TC=TC,TL=TL,DN=DN,TD=TD,TM=TM))

POS=["GK","Thòng","HV","TVtt","TV","TĐ"]
def fit(s,p):
    KT,CH,DD,PN,TC,TL,DN,TD,TM=(s["KT"],s["CH"],s["DD"],s["PN"],s["TC"],s["TL"],s["DN"],s["TD"],s["TM"])
    if p=="GK":   return TM
    if p=="Thòng":return 0.40*PN+0.25*TD+0.20*TL+0.15*CH-0.20*max(0,DD-6)
    if p=="HV":   return 0.30*PN+0.25*TC+0.20*TL+0.15*DN+0.10*TD
    if p=="TVtt": return 0.30*CH+0.30*TD+0.20*KT+0.10*DN+0.10*PN
    if p=="TV":   return 0.30*KT+0.20*TC+0.20*CH+0.15*DN+0.15*DD
    if p=="TĐ":   return 0.45*DD+0.20*KT+0.15*TL+0.10*TD+0.10*TC
OUT=["KT","CH","DD","PN","TC","TL","DN","TD"]
overall=np.array([np.mean([S[i][k] for k in OUT]) for i in range(N)])
TMv=np.array([S[i]["TM"] for i in range(N)])

SLOTS=["GK","Thòng","HV","HV","TVtt","TV","TĐ"]   # doi hinh 1-3-3
# M[i] = gia tri cua cau thu i tren 10 cot [7 slot + 3 bench=0]
# bench=0 -> moi doi luon ra san 7 NGUOI MANH NHAT (best XI), 3 con lai du bi
M=np.zeros((N,10))
for i in range(N):
    for j,sl in enumerate(SLOTS): M[i,j]=fit(S[i],sl)
    M[i,7:]=0.0

# --- Uu tien da chinh: 5 nguoi sau bi DEPRIORITIZE (day ve du bi, tru khi that can) ---
LOWPRIO={"Duy Hùng","Quý Tàu","Trịnh Mạnh","Mạnh Tiến","Quang Minh"}
lowprio=np.array([names[i] in LOWPRIO for i in range(N)])
PEN=3.0   # phat diem cac vi tri NGOAI SAN cua nguoi low-prio (giu nguyen GK -> gon khan hiem)
Mpen=M.copy()
for i in range(N):
    if lowprio[i]: Mpen[i,1:7]-=PEN   # cot 0=GK giu nguyen; 1..6 ngoai san bi phat; 7..9 bench=0

def eval_team(team):
    sub=Mpen[np.ix_(team, range(10))]              # gan vi tri dung diem da phat (uu tien)
    rows,cols=linear_sum_assignment(-sub)          # maximize
    start_fit=0.0; bench_ov=0.0; assign={}; bench=[]
    gk_fit=0.0
    for r_,c_ in zip(rows,cols):
        pl=team[r_]
        if c_<7:
            assign[SLOTS[c_]+("" if SLOTS[c_]!="HV" else f"_{c_}")]=pl
            start_fit+=M[pl,c_]
            if c_==0: gk_fit=M[pl,c_]
        else:
            bench.append(pl); bench_ov+=overall[pl]
    return start_fit, bench_ov, gk_fit, assign, bench

best=None
allidx=list(range(N))
for comb in combinations(range(1,N),9):           # player0 luon o doi A -> khu trung lap
    A=[0]+list(comb); B=[i for i in allidx if i not in A]
    if TMv[A].max()<=0 or TMv[B].max()<=0: continue   # moi doi can it nhat 1 thu mon
    starsA=int(sum(overall[i]>=8 for i in A)); starsB=int(sum(overall[i]>=8 for i in B))
    sA,bA,gA,asA,bnA=eval_team(A)
    sB,bB,gB,asB,bnB=eval_team(B)
    stA=list(asA.values()); stB=list(asB.values())   # 7 da chinh moi doi
    paceA=np.mean([S[p]["TC"] for p in stA]); paceB=np.mean([S[p]["TC"] for p in stB])
    heatA=np.mean([S[p]["DN"] for p in stA]); heatB=np.mean([S[p]["DN"] for p in stB])
    defA=np.mean([S[asA[k]]["PN"] for k in ("Thòng","HV_2","HV_3")])  # PN tuyen duoi
    defB=np.mean([S[asB[k]]["PN"] for k in ("Thòng","HV_2","HV_3")])
    atkA=np.mean([S[asA[k]]["DD"] for k in ("TVtt","TV","TĐ")])       # DĐ tuyen tren
    atkB=np.mean([S[asB[k]]["DD"] for k in ("TVtt","TV","TĐ")])
    J=(3.0*abs(starsA-starsB)            # chia deu ngoi sao (uu tien cao)
       +1.0*abs(sA-sB)                   # can doi hinh chinh
       +0.6*abs((sA+bA)-(sB+bB))         # can tong ca doi
       +0.5*abs(bA-bB)                   # can du bi
       +1.0*abs(defA-defB)              # can PHONG NGU tuyen duoi
       +0.4*abs(atkA-atkB)              # can hang cong tuyen tren
       +0.4*abs(paceA-paceB)             # can toc do
       +0.4*abs(heatA-heatB)             # can do nhiet
       +0.3*abs(gA-gB))                  # can thu mon
    if best is None or J<best[0]:
        best=(J,A,B,(sA,bA,gA,asA,bnA),(sB,bB,gB,asB,bnB))

J,A,B,DA,DB=best
def show(tag,team,D):
    sfit,bov,gfit,asg,bench=D
    print(f"\n=== ĐỘI {tag} === start_fit={sfit:.1f}  bench_ov={bov:.1f}  total={sfit+bov:.1f}  GKfit={gfit:.1f}")
    order=["GK","Thòng","HV_2","HV_3","TVtt","TV","TĐ"]
    lab={"GK":"GK","Thòng":"Thòng","HV_2":"HV","HV_3":"HV","TVtt":"TVtt","TV":"TV","TĐ":"TĐ"}
    for k in order:
        pl=asg[k]; s=S[pl]
        print(f"  {lab[k]:<6} {names[pl]:<19} ov={overall[pl]:.2f}  PN{s['PN']} TC{s['TC']} DĐ{s['DD']} ĐN{s['DN']} TM{s['TM']}")
    print("  DỰ BỊ:", ", ".join(f"{names[p]}(ov{overall[p]:.1f})" for p in sorted(bench,key=lambda p:-overall[p])))
    # metrics
    backs=[asg["Thòng"],asg["HV_2"],asg["HV_3"]]
    fronts=[asg["TVtt"],asg["TV"],asg["TĐ"]]
    defPN=np.mean([S[p]["PN"] for p in backs])
    attk=np.mean([S[p]["DD"] for p in fronts])
    pace=np.mean([S[p]["TC"] for p in (backs+fronts)])
    heat=np.mean([S[p]["DN"] for p in (backs+fronts)])
    stars=sum(1 for p in team if overall[p]>=8)
    return dict(defPN=defPN,attk=attk,pace=pace,heat=heat,stars=stars,start=sfit,bench=bov,total=sfit+bov,gk=gfit)

print(f"BEST J={J:.3f}")
mA=show("A",A,DA); mB=show("B",B,DB)

# ===================== XUAT FILE HTML (so do san bong) =====================
def jersey(kit):
    body,sleeve,trim=kit["body"],kit["sleeve"],kit["trim"]
    return (f'<svg class="jersey" viewBox="0 0 100 90" width="38" height="34">'
            # than ao (body)
            f'<path d="M30,8 L12,18 L4,40 L20,48 L26,40 L26,84 L74,84 L74,40 L80,48 '
            f'L96,40 L88,18 L70,8 C70,8 64,20 50,20 C36,20 30,8 30,8 Z" '
            f'fill="{body}" stroke="#00000055" stroke-width="2"/>'
            # tay ao trai/phai (sleeve)
            f'<path d="M30,8 L12,18 L4,40 L20,48 L26,38 L26,12 Z" fill="{sleeve}" stroke="#00000022" stroke-width="1"/>'
            f'<path d="M70,8 L88,18 L96,40 L80,48 L74,38 L74,12 Z" fill="{sleeve}" stroke="#00000022" stroke-width="1"/>'
            # co ao (trim)
            f'<path d="M34,9 Q50,21 66,9 L62,6 Q50,16 38,6 Z" fill="{trim}"/></svg>')
POSNAME={"GK":"Thủ môn","Thòng":"Trung vệ","HV":"Hậu vệ","TVtt":"Tiền vệ giữa","TV":"Tiền vệ cánh","TĐ":"Tiền đạo"}
def chip(pl, label, kit):
    star=' <span class="star">★</span>' if overall[pl]>=8 else ''
    return (f'<div class="player">{jersey(kit)}'
            f'<div class="name">{names[pl]}{star}</div>'
            f'<div class="sub">{POSNAME.get(label,label)} · {overall[pl]:.1f}</div></div>')
def col(chips, gk=False):
    return f'<div class="col{" gkcol" if gk else ""}">{"".join(chips)}</div>'
def subs_items(bench):
    return ''.join(f'<span class="subchip">{names[p]} <b>{overall[p]:.1f}</b></span>'
                   for p in sorted(bench,key=lambda p:-overall[p]))

CA="#bd2004"; CB="#fcba03"          # header: do Quy Do / vang Phao Thu
CAtxt="#bd2004"; CBtxt="#11224f"    # mau chu dau de cho bao cao
KIT_A={"body":"#bd2004","sleeve":"#911503","trim":"#ffffff"}   # Quy Do (Man Utd) san nha
KIT_B={"body":"#fcba03","sleeve":"#11224f","trim":"#11224f"}   # Phao Thu (Arsenal) san khach
asgA=DA[3]; bnA=DA[4]; asgB=DB[3]; bnB=DB[4]
# cot cau thu cho san NGANG: A (trai, tan cong phai) | B (phai, tan cong trai)
cAgk =col([chip(asgA["GK"],"GK",KIT_A)],True)
cAdef=col([chip(asgA["HV_2"],"HV",KIT_A),chip(asgA["Thòng"],"Thòng",KIT_A),chip(asgA["HV_3"],"HV",KIT_A)])
cAatt=col([chip(asgA["TVtt"],"TVtt",KIT_A),chip(asgA["TĐ"],"TĐ",KIT_A),chip(asgA["TV"],"TV",KIT_A)])
cBatt=col([chip(asgB["TVtt"],"TVtt",KIT_B),chip(asgB["TĐ"],"TĐ",KIT_B),chip(asgB["TV"],"TV",KIT_B)])
cBdef=col([chip(asgB["HV_2"],"HV",KIT_B),chip(asgB["Thòng"],"Thòng",KIT_B),chip(asgB["HV_3"],"HV",KIT_B)])
cBgk =col([chip(asgB["GK"],"GK",KIT_B)],True)
faster="A" if mA['pace']>mB['pace'] else "B"; slower="B" if faster=="A" else "A"
dtot=abs(mA['total']-mB['total'])
pred=("Trận 50-50 — chênh tổng chỉ ~%.1f%%. Dự đoán <b>3-3</b> hoặc chênh 1 bàn; "
      "đội %s nhanh hơn nên nhỉnh nhẹ ở phản công, đội %s kiểm soát tốt thì thắng sít sao."
      % (100*dtot/((mA['total']+mB['total'])/2), faster, slower))
def card(m, o, ctxt, label, asg, bench):
    fast = m['pace']>o['pace']
    nm=lambda k: names[asg[k]]
    slots=["GK","Thòng","HV_2","HV_3","TVtt","TV","TĐ"]
    stars=[names[asg[k]] for k in slots if overall[asg[k]]>=8]
    topbench=names[max(bench,key=lambda p:overall[p])] if bench else "—"
    b=[]
    if stars:
        b.append(f"⭐ Đầu tàu <b>{' + '.join(stars)}</b> — gánh điểm và tạo đột biến ở mọi thời điểm.")
    b.append(f"🎯 Hàng công sắc, dứt điểm TB <b>{m['attk']:.1f}/10</b>, mũi nhọn <b>{nm('TĐ')}</b> chốt hạ trước khung thành.")
    b.append(f"🛡️ Tuyến dưới chắc, phòng ngự TB <b>{m['defPN']:.1f}/10</b>: thòng <b>{nm('Thòng')}</b> bọc lót, thủ môn <b>{nm('GK')}</b> trấn giữ.")
    if fast:
        b.append(f"⚡ Tốc độ vượt trội (<b>{m['pace']:.1f}/10</b>) → phản công chớp nhoáng, pressing tầm cao ngạt thở.")
    else:
        b.append(f"🧠 Đồng đều & lì đòn (độ nhiệt <b>{m['heat']:.1f}/10</b>) → cầm bóng, đập nhả, áp đặt thế trận bằng kỷ luật.")
    b.append(f"🔁 Băng ghế còn <b>{topbench}</b> để bơm sức, giữ cường độ tới phút chót.")
    weak=("khi mất bóng cần 3 mũi trên tích cực lùi về; gặp đối thủ biết cầm nhịp sẽ phải kiên nhẫn hơn"
          if fast else
          "tốc độ thấp hơn một nhịp — phải bọc lót kín, tránh để đối thủ thốc thẳng sau lưng")
    plan=("đẩy cao pressing, đoạt bóng nhanh rồi tung tốc độ kết liễu — ghi bàn sớm để dẫn dắt thế trận"
          if fast else
          "kiểm soát bóng, kéo giãn đối thủ rồi phối hợp trung lộ tung đòn quyết định")
    lis=''.join(f'<li>{x}</li>' for x in b)
    return (f'<div class="card"><h3 style="color:{ctxt}">{label}</h3>'
            f'<ul class="ul">{lis}</ul>'
            f'<div class="wk">⚠️ <b>Điểm cần lưu ý:</b> {weak}.</div>'
            f'<div class="pl">🏆 <b>Cách thắng:</b> {plan}.</div></div>')
cardA=card(mA,mB,CAtxt,"Đội A · Quỷ Đỏ (sân nhà)",asgA,bnA)
cardB=card(mB,mA,CBtxt,"Đội B · Pháo Thủ (sân khách)",asgB,bnB)
balintro=("Hai đội chênh tổng chỉ <b>~%.1f%%</b> — gần như <b>50/50</b>. Khác biệt nằm ở <b>chất chơi</b> chứ không phải trình độ: "
          "một bên thiên <b>tốc độ – phản công</b>, một bên thiên <b>kiểm soát – chắc chắn</b>. Mỗi đội đều có 2 ngôi sao, "
          "thủ môn ngang nhau, hàng công và hàng thủ tương đương — nên đội nào cũng có vũ khí riêng và cửa thắng sòng phẳng. Ra sân cứ tự tin!"
          % (100*dtot/((mA['total']+mB['total'])/2)))

def metric_rows():
    R=[("Sức mạnh đội hình chính",f"{mA['start']:.1f}",f"{mB['start']:.1f}"),
       ("Tổng cả đội (chính + dự bị)",f"{mA['total']:.1f}",f"{mB['total']:.1f}"),
       ("Ngôi sao (điểm tổng ≥ 8)",f"{mA['stars']}",f"{mB['stars']}"),
       ("Thủ môn",f"{mA['gk']:.1f}",f"{mB['gk']:.1f}"),
       ("Phòng ngự tuyến dưới",f"{mA['defPN']:.1f}",f"{mB['defPN']:.1f}"),
       ("Dứt điểm tuyến trên",f"{mA['attk']:.1f}",f"{mB['attk']:.1f}"),
       ("Tốc độ trung bình",f"{mA['pace']:.1f}",f"{mB['pace']:.1f}"),
       ("Độ nhiệt (chịu chạy)",f"{mA['heat']:.1f}",f"{mB['heat']:.1f}")]
    return ''.join(f'<tr><td>{n}</td><td>{a}</td><td>{b}</td></tr>' for n,a,b in R)

CSS="""
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:#eef1f6;color:#1c2533;padding:24px}
h1{text-align:center;font-size:22px;margin-bottom:4px;color:#16203a}
.note{text-align:center;color:#5a6a82;font-size:13px;margin-bottom:22px}
.match{max-width:832px;margin:0 auto}
.mhead{display:flex;border-radius:8px 8px 0 0;overflow:hidden}
.mh{flex:1;font-weight:800;font-size:13px;padding:8px 12px;letter-spacing:.3px}
.mh.l{text-align:left}.mh.r{text-align:right}
.field{position:relative;height:384px;overflow:hidden;
 background:repeating-linear-gradient(90deg,#46b657 0 58px,#3da94c 58px 116px);
 border:1px solid #cfd6e2;border-top:0;display:flex;align-items:stretch;padding:6px 2px;box-shadow:0 3px 12px #0000001f}
.mk{position:absolute;border:2px solid #ffffff9a}
.vhalf{left:50%;top:8px;bottom:8px;width:0;border:0;border-left:2px solid #ffffff9a}
.ccircle{width:80px;height:80px;border-radius:50%;left:50%;top:50%;transform:translate(-50%,-50%)}
.boxL{left:0;top:50%;transform:translateY(-50%);width:46px;height:176px;border-left:0}
.boxR{right:0;top:50%;transform:translateY(-50%);width:46px;height:176px;border-right:0}
.goalL{left:0;top:50%;transform:translateY(-50%);width:8px;height:55px;background:#ffffff55;border:0}
.goalR{right:0;top:50%;transform:translateY(-50%);width:8px;height:55px;background:#ffffff55;border:0}
.col{flex:1;display:flex;flex-direction:column;justify-content:space-around;align-items:center;position:relative;z-index:2}
.gkcol{flex:.72}
.player{display:flex;flex-direction:column;align-items:center;width:84px}
.jersey{filter:drop-shadow(0 2px 2px #0005)}
.name{color:#fff;font-weight:700;font-size:10px;margin-top:1px;text-align:center;line-height:1.1;text-shadow:0 1px 3px #000b}
.sub{color:#f4fff7;font-size:8px;margin-top:0;text-shadow:0 1px 3px #000b;opacity:.95}
.star{color:#ffe14d}
.benchrow{display:flex;border-radius:0 0 8px 8px;overflow:hidden}
.benchcol{flex:1;background:#fff;padding:8px 11px;font-size:11px;color:#33415c;border:1px solid #e2e7f0;border-top:0}
.benchcol.l{border-right:0}
.subs-lbl{color:#5a6a82;font-weight:700;margin-right:4px}
.subchip{display:inline-block;background:#f2f5fa;border:1px solid #d7deea;padding:2px 7px;border-radius:20px;color:#26314a;margin:2px 3px;font-size:11px}
.report{max-width:820px;margin:30px auto 0}
.report h2{font-size:16px;margin:18px 0 8px;color:#1d2b4a}
table{width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px #00000014}
th,td{padding:8px 12px;border-bottom:1px solid #e6eaf1;text-align:center}
td:first-child{text-align:left;color:#33415c}
th{background:#1d2b4a;color:#fff}
.cards{display:flex;gap:16px;flex-wrap:wrap;margin-top:6px}
.card{flex:1;min-width:260px;background:#fff;border:1px solid #e2e7f0;border-radius:8px;padding:12px 14px;font-size:13px;line-height:1.5;box-shadow:0 2px 8px #00000012}
.card h3{font-size:14px;margin-bottom:6px}
.card .ul{margin:2px 0 0;padding-left:19px}
.card li{margin:4px 0;line-height:1.45}
.wk{margin-top:8px;padding-top:7px;border-top:1px dashed #e6eaf1;color:#8a5a00;font-size:12.5px;line-height:1.45}
.pl{margin-top:6px;color:#15602a;font-size:12.5px;line-height:1.45}
.persuade{background:#fff;border:1px solid #e2e7f0;border-left:4px solid #1d2b4a;border-radius:8px;padding:11px 14px;margin:2px 0 12px;font-size:13.5px;line-height:1.55;color:#33415c}
.pred{background:#eaf6ec;border:1px solid #bfe0c5;border-radius:8px;padding:11px 16px;margin:0 auto 16px;max-width:832px;text-align:center;font-size:14px;line-height:1.5;color:#1c3a23}
.foot{text-align:center;color:#8a96a8;font-size:12px;margin-top:22px}
"""
HTML=f"""<!doctype html><html lang="vi"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Đội hình 1-3-3 — 2 đội cân bằng</title><style>{CSS}</style></head><body>
<h1>Đội hình thi đấu</h1>
<div class="note">Duyệt toàn bộ 92.378 cách chia · chênh tổng ~{100*dtot/((mA['total']+mB['total'])/2):.1f}%</div>
<div class="pred">🔮 <b>Dự đoán:</b> {pred}</div>
<div class="match">
 <div class="mhead">
  <div class="mh l" style="background:{CA};color:#fff">🔴 ĐỘI A · Quỷ Đỏ (sân nhà) · Tổng {mA['total']:.1f} ▶</div>
  <div class="mh r" style="background:{CB};color:#11224f">◀ Tổng {mB['total']:.1f} · Pháo Thủ (sân khách) · ĐỘI B 🟡</div>
 </div>
 <div class="field">
  <div class="mk vhalf"></div><div class="mk ccircle"></div>
  <div class="mk boxL"></div><div class="mk boxR"></div>
  <div class="mk goalL"></div><div class="mk goalR"></div>
  {cAgk}{cAdef}{cAatt}{cBatt}{cBdef}{cBgk}
 </div>
 <div class="benchrow">
  <div class="benchcol l"><span class="subs-lbl">Dự bị A:</span>{subs_items(bnA)}</div>
  <div class="benchcol"><span class="subs-lbl">Dự bị B:</span>{subs_items(bnB)}</div>
 </div>
</div>
<div class="report">
 <h2>⚖️ Ưu – nhược điểm</h2>
 <div class="persuade">{balintro}</div>
 <div class="cards">{cardA}{cardB}</div>
 <h2>📊 Chỉ số đánh giá</h2>
 <table><tr><th>Chỉ số</th><th>Đội A</th><th>Đội B</th></tr>{metric_rows()}</table>
</div>
<div class="foot">Tạo tự động từ bảng đánh giá · sơ đồ 1-3-3 (1 thủ môn · 1 trung vệ + 2 hậu vệ · tiền vệ giữa + tiền vệ cánh + tiền đạo)</div>
</body></html>"""
outhtml="/Users/tuyennd/Documents/VN_NAMI/Tools/GenPlayer/preview_static.html"   # ban Python (tinh, de doi chieu) - trang web that la index.html do build_web.py tao
with open(outhtml,"w",encoding="utf-8") as f: f.write(HTML)
print("\nĐã xuất HTML:", outhtml)
print("\n--- CHENH LECH ---")
print(f"Start XI : A={mA['start']:.1f}  B={mB['start']:.1f}  | lệch {abs(mA['start']-mB['start']):.2f}")
print(f"Dự bị    : A={mA['bench']:.1f}  B={mB['bench']:.1f}  | lệch {abs(mA['bench']-mB['bench']):.2f}")
print(f"Tổng     : A={mA['total']:.1f}  B={mB['total']:.1f}  | lệch {abs(mA['total']-mB['total']):.2f}")
print(f"Thủ môn  : A={mA['gk']:.1f}  B={mB['gk']:.1f}")
print(f"PN tuyến dưới: A={mA['defPN']:.1f} B={mB['defPN']:.1f} | Dứt điểm tuyến trên: A={mA['attk']:.1f} B={mB['attk']:.1f}")
print(f"Tốc độ TB: A={mA['pace']:.1f} B={mB['pace']:.1f} | Độ nhiệt TB: A={mA['heat']:.1f} B={mB['heat']:.1f} | Sao(>=8): A={mA['stars']} B={mB['stars']}")
